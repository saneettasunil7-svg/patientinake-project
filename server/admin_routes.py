from datetime import datetime
import shutil
import os
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from typing import List
from sqlalchemy.orm import Session
from database import get_db
import models
import appointment_models
import admin_schemas
import audit_service
import auth

router = APIRouter(prefix="/admin", tags=["admin"])

UPLOAD_DIR = "uploads/profiles"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

def verify_admin(current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

@router.post("/doctors", response_model=admin_schemas.DoctorResponse)
async def create_doctor(
    doctor: admin_schemas.DoctorCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    # Check if email exists
    existing_user = db.query(models.User).filter(models.User.email == doctor.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user account
    hashed_password = auth.get_password_hash(doctor.password)
    new_user = models.User(
        email=doctor.email,
        hashed_password=hashed_password,
        role="doctor",
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Create doctor profile
    doctor_profile = models.DoctorProfile(
        user_id=new_user.id,
        full_name=doctor.full_name,
        specialization=doctor.specialization,
        bio=doctor.bio
    )
    db.add(doctor_profile)
    
    # Create availability record
    availability = appointment_models.DoctorAvailability(
        doctor_id=new_user.id,
        is_available=False
    )
    db.add(availability)
    db.commit()
    db.refresh(doctor_profile)
    
    return {
        "id": new_user.id,
        "email": new_user.email,
        "full_name": doctor_profile.full_name,
        "specialization": doctor_profile.specialization,
        "bio": doctor_profile.bio,
        "profile_photo": doctor_profile.profile_photo,
        "is_active": new_user.is_active,
        "is_verified": doctor_profile.is_verified,
        "is_available": False
    }

@router.get("/doctors", response_model=List[admin_schemas.DoctorResponse])
async def list_doctors(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctors = db.query(models.User).filter(models.User.role == "doctor").all()
    result = []
    for doctor in doctors:
        if doctor.doctor_profile:
            availability = db.query(appointment_models.DoctorAvailability).filter(
                appointment_models.DoctorAvailability.doctor_id == doctor.id
            ).first()
            result.append({
                "id": doctor.id,
                "email": doctor.email,
                "full_name": doctor.doctor_profile.full_name,
                "specialization": doctor.doctor_profile.specialization,
                "bio": doctor.doctor_profile.bio,
                "profile_photo": doctor.doctor_profile.profile_photo,
                "is_active": doctor.is_active,
                "is_verified": doctor.doctor_profile.is_verified,
                "is_available": availability.is_available if availability else False
            })
    return result

@router.get("/doctors/{doctor_id}", response_model=admin_schemas.DoctorResponse)
async def get_doctor(
    doctor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctor = db.query(models.User).filter(
        models.User.id == doctor_id,
        models.User.role == "doctor"
    ).first()
    
    if not doctor or not doctor.doctor_profile:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    availability = db.query(appointment_models.DoctorAvailability).filter(
        appointment_models.DoctorAvailability.doctor_id == doctor.id
    ).first()
    
    return {
        "id": doctor.id,
        "email": doctor.email,
        "full_name": doctor.doctor_profile.full_name,
        "specialization": doctor.doctor_profile.specialization,
        "bio": doctor.doctor_profile.bio,
        "profile_photo": doctor.doctor_profile.profile_photo,
        "is_active": doctor.is_active,
        "is_available": availability.is_available if availability else False
    }

@router.put("/doctors/{doctor_id}", response_model=admin_schemas.DoctorResponse)
async def update_doctor(
    doctor_id: int,
    doctor_update: admin_schemas.DoctorUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctor = db.query(models.User).filter(
        models.User.id == doctor_id,
        models.User.role == "doctor"
    ).first()
    
    if not doctor or not doctor.doctor_profile:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    # Update doctor profile fields
    if doctor_update.full_name is not None:
        doctor.doctor_profile.full_name = doctor_update.full_name
    if doctor_update.specialization is not None:
        doctor.doctor_profile.specialization = doctor_update.specialization
    if doctor_update.bio is not None:
        doctor.doctor_profile.bio = doctor_update.bio
    if doctor_update.is_verified is not None:
        doctor.doctor_profile.is_verified = doctor_update.is_verified
    
    # Update user fields
    if doctor_update.is_active is not None:
        doctor.is_active = doctor_update.is_active

    db.commit()
    db.refresh(doctor)
    
    # Log action
    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "UPDATE_DOCTOR", 
        f"Updated doctor profile for {doctor.email}"
    )
    
    availability = db.query(appointment_models.DoctorAvailability).filter(
        appointment_models.DoctorAvailability.doctor_id == doctor.id
    ).first()
    
    return {
        "id": doctor.id,
        "email": doctor.email,
        "full_name": doctor.doctor_profile.full_name,
        "specialization": doctor.doctor_profile.specialization,
        "bio": doctor.doctor_profile.bio,
        "profile_photo": doctor.doctor_profile.profile_photo,
        "is_active": doctor.is_active,
        "is_verified": doctor.doctor_profile.is_verified,
        "is_available": availability.is_available if availability else False
    }

@router.delete("/doctors/{doctor_id}")
async def delete_doctor(
    doctor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctor = db.query(models.User).filter(
        models.User.id == doctor_id,
        models.User.role == "doctor"
    ).first()
    
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    email = doctor.email
    
    # Hard delete related records first
    db.query(models.DoctorProfile).filter(models.DoctorProfile.user_id == doctor_id).delete()
    db.query(appointment_models.DoctorAvailability).filter(appointment_models.DoctorAvailability.doctor_id == doctor_id).delete()
    db.query(appointment_models.DoctorSchedule).filter(appointment_models.DoctorSchedule.doctor_id == doctor_id).delete()
    
    # Finally delete the user
    db.delete(doctor)
    db.commit()

    # Log action
    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "DELETE_DOCTOR", 
        f"Hard deleted doctor and profile: {email}"
    )
    
    return {"message": "Doctor deleted successfully"}

@router.put("/doctors/{doctor_id}/verify")
async def verify_doctor(
    doctor_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctor = db.query(models.User).filter(
        models.User.id == doctor_id,
        models.User.role == "doctor"
    ).first()
    
    if not doctor or not doctor.doctor_profile:
        raise HTTPException(status_code=404, detail="Doctor not found")
        
    doctor.doctor_profile.is_verified = True
    db.commit()

    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "VERIFY_DOCTOR", 
        f"Verified doctor {doctor.email}"
    )

    return {"message": "Doctor verified successfully"}

@router.post("/doctors/{doctor_id}/photo", response_model=admin_schemas.DoctorResponse)
async def upload_doctor_photo(
    doctor_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    doctor = db.query(models.User).filter(
        models.User.id == doctor_id,
        models.User.role == "doctor"
    ).first()
    
    if not doctor or not doctor.doctor_profile:
        raise HTTPException(status_code=404, detail="Doctor not found")
        
    # Generate unique filename
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    extension = os.path.splitext(file.filename)[1]
    filename = f"profile_{doctor.id}_{timestamp}{extension}"
    file_path = os.path.join(UPLOAD_DIR, filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_url = f"/static/profiles/{filename}"
    
    # Update user model
    doctor.profile_photo = photo_url
    
    # Update associated profile
    doctor.doctor_profile.profile_photo = photo_url
            
    db.commit()
    db.refresh(doctor)
    
    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "UPLOAD_DOCTOR_PHOTO", 
        f"Uploaded photo for doctor {doctor.email}"
    )
    
    availability = db.query(appointment_models.DoctorAvailability).filter(
        appointment_models.DoctorAvailability.doctor_id == doctor.id
    ).first()
    
    return {
        "id": doctor.id,
        "email": doctor.email,
        "full_name": doctor.doctor_profile.full_name,
        "specialization": doctor.doctor_profile.specialization,
        "bio": doctor.doctor_profile.bio,
        "profile_photo": doctor.doctor_profile.profile_photo,
        "is_active": doctor.is_active,
        "is_verified": doctor.doctor_profile.is_verified,
        "is_available": availability.is_available if availability else False
    }

# --- Policies ---

@router.get("/policies", response_model=List[admin_schemas.PolicyResponse])
async def list_policies(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    return db.query(models.Policy).all()

@router.get("/policies/public", response_model=List[admin_schemas.PolicyResponse])
async def list_public_policies(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    return db.query(models.Policy).all()

@router.post("/policies", response_model=admin_schemas.PolicyResponse)
async def create_policy(
    policy: admin_schemas.PolicyCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    new_policy = models.Policy(
        title=policy.title,
        content=policy.content,
        last_updated=datetime.utcnow().isoformat()
    )
    db.add(new_policy)
    db.commit()
    db.refresh(new_policy)

    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "CREATE_POLICY", 
        f"Created policy {policy.title}"
    )
    return new_policy

@router.put("/policies/{policy_id}", response_model=admin_schemas.PolicyResponse)
async def update_policy(
    policy_id: int,
    policy_update: admin_schemas.PolicyCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    existing_policy = db.query(models.Policy).filter(models.Policy.id == policy_id).first()
    if not existing_policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    
    existing_policy.title = policy_update.title
    existing_policy.content = policy_update.content
    existing_policy.last_updated = datetime.utcnow().isoformat()
    
    db.commit()
    db.refresh(existing_policy)
    
    audit_service.log_admin_action(
        db, 
        current_user.id, 
        "UPDATE_POLICY", 
        f"Updated policy {existing_policy.title}"
    )
    return existing_policy

# --- Audit Logs ---

@router.get("/audit-logs", response_model=List[admin_schemas.AuditLogResponse])
async def get_audit_logs(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    return db.query(models.AuditLog).order_by(models.AuditLog.id.desc()).limit(100).all()

# --- Reports ---

@router.get("/reports", response_model=admin_schemas.StatsResponse)
async def get_reports(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    total_doctors = db.query(models.User).filter(models.User.role == "doctor").count()
    active_doctors = db.query(models.User).filter(models.User.role == "doctor", models.User.is_active == True).count()
    verified_doctors = db.query(models.DoctorProfile).filter(models.DoctorProfile.is_verified == True).count()
    total_patients = db.query(models.User).filter(models.User.role == "patient").count()
    total_appointments = db.query(appointment_models.Appointment).count()

    return {
        "total_doctors": total_doctors,
        "active_doctors": active_doctors,
        "verified_doctors": verified_doctors,
        "total_patients": total_patients,
        "total_appointments": total_appointments
    }

@router.get("/appointments")
async def get_all_appointments(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    appointments = db.query(appointment_models.Appointment).all()
    return appointments

@router.get("/export/appointments")
async def export_appointments_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(verify_admin)
):
    """Export all token/appointment records to an Excel file."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    # ------ Header Row ------
    headers = [
        "Token #", "Patient Name", "Patient Email",
        "Doctor Name", "Department", "Status",
        "Date Created", "Reason for Visit", "Is Emergency"
    ]
    header_fill = PatternFill("solid", fgColor="1E40AF")   # deep blue
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # Column widths
    col_widths = [10, 22, 28, 22, 18, 14, 22, 35, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ------ Data Rows ------
    tokens = db.query(appointment_models.Token).order_by(
        appointment_models.Token.created_at.desc()
    ).all()

    alt_fill = PatternFill("solid", fgColor="EFF6FF")  # light blue-gray for alternating rows

    for row_idx, token in enumerate(tokens, start=2):
        patient = db.query(models.User).filter(models.User.id == token.patient_id).first()
        patient_profile = db.query(models.PatientProfile).filter(
            models.PatientProfile.user_id == token.patient_id
        ).first()
        doctor_profile = db.query(models.DoctorProfile).filter(
            models.DoctorProfile.user_id == token.doctor_id
        ).first() if token.doctor_id else None

        row_data = [
            token.token_number,
            patient_profile.full_name if patient_profile else (patient.email.split('@')[0] if patient else "Unknown"),
            patient.email if patient else "Unknown",
            doctor_profile.full_name if doctor_profile else "Unassigned",
            doctor_profile.specialization if doctor_profile else "N/A",
            token.status.replace("_", " ").title(),
            token.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(token.created_at, "strftime") else str(token.created_at)[:16],
            token.reason_for_visit or "",
            "Yes" if token.is_emergency else "No"
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Stream the file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"appointments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
